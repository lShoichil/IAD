import copy
import itertools
import random
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.workbook import Workbook

COUNT_CLASS: int = 2
COUNT_FEATURE_INTO_CLASS: int = 6

# TODO: Может упасть если COUNT_CATEGORICAL_FEATURES_START == 1
COUNT_CATEGORICAL_FEATURES_START: int = 6
COUNT_CATEGORICAL_FEATURES_END: int = 12

COUNT_CHPD_START: int = 2
COUNT_CHPD_END: int = 3

LOWER_BOUND: int = 2
UPPER_BOUND: int = 24

COUNT_MEDICAL_HISTORY_FOR_ONE_CLASS: int = 30
COUNT_OBS_MOMENT_START: int = 1
COUNT_OBS_MOMENT_END: int = 3

# ----------------------------------------------- Расчёты -----------------------------------------------
# ----------------------------------------------- МБЗ -----------------------------------------------
classesList = [f'Класс {i}' for i in range(1, COUNT_CLASS + 1)]
featuresList = [f'Признак {i}' for i in range(1, COUNT_FEATURE_INTO_CLASS + 1)]
typesFeaturesMap = {'Перечислимый': [], 'Категориальный': [], 'Бинарный': []}

for i, feature in enumerate(featuresList):
    if i % 3 == 0:
        typesFeaturesMap['Перечислимый'].append(feature)
    elif i % 3 == 1:
        typesFeaturesMap['Категориальный'].append(feature)
    else:
        typesFeaturesMap['Бинарный'].append(feature)
    if i + 1 >= COUNT_FEATURE_INTO_CLASS:
        break

# {'Класс1': {'Признак1': 2, 'Признак2': 4, ...}, 'Класс2': {...}}
amountPeriodDynamicsForFeaturesInClassMap = {}
for class_name in classesList:
    amountPeriodDynamicsForFeaturesMap = {
        feature: random.randint(COUNT_CHPD_START, COUNT_CHPD_END) for feature in featuresList
    }
    amountPeriodDynamicsForFeaturesInClassMap[class_name] = amountPeriodDynamicsForFeaturesMap
invertedTypesFeaturesMap = {v: k for k, features in typesFeaturesMap.items() for v in features}

# {'Признак1': (1, 4), 'Признак2': ['v0', 'v1', 'v2', 'v3', 'v4'], 'Признак3': [0, 1], ... }
possibleValuesForFeaturesMap = {}
for feature in featuresList:
    feature_type = invertedTypesFeaturesMap[feature]
    if feature_type == 'Перечислимый':
        # Генерируем отрезок вида [a,b], где a,b - числа > 0, a > b и
        # разница между ними >= value признака из amountPeriodDynamicsForFeaturesInClassMap
        value = amountPeriodDynamicsForFeaturesInClassMap[classesList[0]][
            feature]  # Для упрощения берем значение из первого класса
        a: int = 0
        b: int = 0
        while a - value < 10:
            a = random.randint(value, 10 * value)
        b = random.randint(1, a - value)
        possibleValuesForFeaturesMap[feature] = (b, a + COUNT_CHPD_END + 150)
    elif feature_type == 'Категориальный':
        n = random.randint(COUNT_CATEGORICAL_FEATURES_START, COUNT_CATEGORICAL_FEATURES_END)
        possibleValuesForFeaturesMap[feature] = [f'v{i}' for i in range(n)]
    elif feature_type == 'Бинарный':
        # Всегда генерируем list из значения 0 и 1
        possibleValuesForFeaturesMap[feature] = [0, 1]

# Создаём Map для ЗДП
# {'Признак1': {1: есть, 2: нет}, 'Признак2': ...}
improvedMap = {}
# {'Класс1': {'Признак1': {1: [], 2: [], 3: [], 4: [], 5: []}, ... }
improvedAmountPeriodDynamicsForFeaturesInClassMap = {}
for class_name in classesList:
    improvedAmountPeriodDynamicsForFeaturesMap = {}
    for feature, dynamicsMap in amountPeriodDynamicsForFeaturesInClassMap[class_name].items():
        improvedMap[feature] = {}
        possibleValues = copy.deepcopy(possibleValuesForFeaturesMap[feature])
        if feature in typesFeaturesMap['Категориальный']:
            for i in range(dynamicsMap):
                if i == 0:
                    improvedMap[feature][i + 1] = random.sample(possibleValues, COUNT_CATEGORICAL_FEATURES_START // 2)
                else:
                    improvedMap[feature][i + 1] = [value for value in possibleValues[:random.randint(COUNT_CATEGORICAL_FEATURES_START, len(possibleValues))] if value not in improvedMap[feature][i]]
        elif feature in typesFeaturesMap['Бинарный']:
            for i in range(dynamicsMap):
                if i % 2 == 0:
                    improvedMap[feature][i + 1] = possibleValues[0]
                else:
                    improvedMap[feature][i + 1] = possibleValues[1]
        elif feature in typesFeaturesMap['Перечислимый']:
            a, b = possibleValues  # входной отрезок
            segment_length = b - a  # длина отрезка
            subsegment_length = segment_length / n  # длина каждого подотрезка
            for i in range(1,dynamicsMap + 1):
                # вычисляем границы текущего подотрезка
                start = a + i * subsegment_length
                end = a + (i + 1) * subsegment_length
                # округляем границы до целых чисел
                start = round(start)
                end = round(end)
                # последний подотрезок может быть немного короче, чтобы компенсировать погрешности округления
                if i == n - 1:
                    end = b
                # использовать подотрезок в работе
                subsegment = (start, end - 1)
                improvedMap[feature][i] = subsegment
        improvedAmountPeriodDynamicsForFeaturesMap[feature] = improvedMap[feature]
    improvedAmountPeriodDynamicsForFeaturesInClassMap[class_name] = improvedAmountPeriodDynamicsForFeaturesMap

# print(possibleValuesForFeaturesMap)
# print(improvedMap)
# print(improvedAmountPeriodDynamicsForFeaturesInClassMap)

# Создаём Map для Верхние и нижние границы (ВГ и НГ)
# {'Признак1': {1: (12, 24), 2: (3, 13), 3: (10, 12)}, ...
improvedVGNGMap = {}
# {'Класс1': {'Признак1': {1: (12, 24), 2: (3, 13), ...}}
improvedAmountPeriodDynamicsForFeaturesInClassVGNGMap = {}
for class_name in classesList:
    improvedAmountPeriodDynamicsForFeaturesMap = {}
    for feature, dynamicsMap in amountPeriodDynamicsForFeaturesInClassMap[class_name].items():
        improvedVGNGMap[feature] = {}
        for i in range(1, dynamicsMap + 1):
            a = random.randint(LOWER_BOUND, UPPER_BOUND)
            b = random.randint(LOWER_BOUND, UPPER_BOUND)
            while a >= b:
                a = random.randint(LOWER_BOUND, UPPER_BOUND)
                b = random.randint(LOWER_BOUND, UPPER_BOUND)
            improvedVGNGMap[feature][i] = (a, b)
        improvedAmountPeriodDynamicsForFeaturesMap[feature] = improvedVGNGMap[feature]
    improvedAmountPeriodDynamicsForFeaturesInClassVGNGMap[class_name] = improvedAmountPeriodDynamicsForFeaturesMap

# print(improvedVGNGMap)
# print(improvedAmountPeriodDynamicsForFeaturesInClassVGNGMap)

# ----------------------------------------------- МВД -----------------------------------------------

medicalHistoryList = [f'ИБ {i}' for i in range(1, (COUNT_MEDICAL_HISTORY_FOR_ONE_CLASS) + 1)]

# (ИБ, заболевание, признак, номер ПД, длительность ПД, число МН в ПД)
# {'ИБ 1': {'Класс 1': {'Признак 1': {1: (16, 3), 2: (14, 2)}, где
# () - (DURATION_DYNAMICS_PERIOD, COUNT_OBS_MOMENT_IN_DYNAMIC_PERIOD)
medicalHistoryMap = {}
for medical_history in medicalHistoryList:
    medicalHistoryMap[medical_history] = {}
    for class_name in classesList:
        medicalHistoryMap[medical_history][class_name] = {}
        for feature, dynamicsMap in amountPeriodDynamicsForFeaturesInClassMap[class_name].items():
            medicalHistoryMap[medical_history][class_name][feature] = {}
            for i in range(1, dynamicsMap + 1):
                PD = improvedAmountPeriodDynamicsForFeaturesInClassVGNGMap[class_name][feature][i]
                DURATION_DYNAMICS_PERIOD = random.randint(PD[0], PD[1])
                if DURATION_DYNAMICS_PERIOD < COUNT_OBS_MOMENT_END:
                    COUNT_OBS_MOMENT_IN_DYNAMIC_PERIOD = random.randint(COUNT_OBS_MOMENT_START,
                                                                        DURATION_DYNAMICS_PERIOD)
                else:
                    COUNT_OBS_MOMENT_IN_DYNAMIC_PERIOD = random.randint(COUNT_OBS_MOMENT_START, COUNT_OBS_MOMENT_END)
                medicalHistoryMap[medical_history][class_name][feature][i] = {
                    'Длительность ПД': DURATION_DYNAMICS_PERIOD, 'Число МН в ПД': COUNT_OBS_MOMENT_IN_DYNAMIC_PERIOD}

# Выборка данных (ИБ, заболевание, признак, МН, значение в МН)
data_sampling = {}
count = 0
for medical_history in medicalHistoryList:
    data_sampling[medical_history] = {}
    for class_name in classesList:
        data_sampling[medical_history][class_name] = {}
        for feature in medicalHistoryMap[medical_history][class_name]:
            data_sampling[medical_history][class_name][feature] = {}
            generate_obs_moment: int = 0
            prev_generated_obs_moment: int = 0
            for j, value in enumerate(medicalHistoryMap[medical_history][class_name][feature].values()):
                duration_dynamic_period_from_medicalHistoryMap = value['Длительность ПД']
                for i in range(1, value['Число МН в ПД'] + 1):
                    generate_obs_moment_value = None
                    # ----------------------------------------------------------------
                    # todo: fix possible crash loop
                    while generate_obs_moment == prev_generated_obs_moment:
                        if len(data_sampling[medical_history][class_name][feature]) < \
                                medicalHistoryMap[medical_history][class_name][feature][1]['Число МН в ПД']:
                            generate_obs_moment = random.randint(1, duration_dynamic_period_from_medicalHistoryMap - 1)
                        else:
                            generate_obs_moment = random.randint(prev_generated_obs_moment + 1,
                                                                 prev_generated_obs_moment + duration_dynamic_period_from_medicalHistoryMap - 1)
                    # ----------------------------------------------------------------
                    ZDP = improvedAmountPeriodDynamicsForFeaturesInClassMap[class_name][feature][j + 1]
                    if feature in typesFeaturesMap['Перечислимый']:
                        left = ZDP[0]
                        right = ZDP[1]
                        generate_obs_moment_value = random.randint(left, right)
                    elif feature in typesFeaturesMap['Категориальный']:
                        generate_obs_moment_value = random.choice(ZDP)
                    elif feature in typesFeaturesMap['Бинарный']:
                        generate_obs_moment_value = ZDP
                    # ----------------------------------------------------------------
                    prev_generated_obs_moment = generate_obs_moment
                    count += 1
                    data_sampling[medical_history][class_name][feature][generate_obs_moment] = generate_obs_moment_value
                    # ----------------------------------------------------------------
            sorted_data_sampling_feature = {k: v for k, v in
                                            sorted(data_sampling[medical_history][class_name][feature].items(),
                                                   key=lambda item: item[0])}
            data_sampling[medical_history][class_name][feature] = sorted_data_sampling_feature

print('Nope crash loop')
# print(medicalHistoryMap)
# print(data_sampling)
# ----------------------------------------------- ИФБЗ -----------------------------------------------
first_column = {}
for medical_history in data_sampling:
    for class_name in data_sampling[medical_history]:
        for feature in data_sampling[medical_history][class_name]:
            for generate_obs_moment in data_sampling[medical_history][class_name][feature]:
                value = data_sampling[medical_history][class_name][feature][generate_obs_moment]
                if class_name not in first_column:
                    first_column[class_name] = {}
                if feature not in first_column[class_name]:
                    first_column[class_name][feature] = {}
                if medical_history not in first_column[class_name][feature]:
                    first_column[class_name][feature][medical_history] = {}
                first_column[class_name][feature][medical_history][generate_obs_moment] = value
# print(f'first_column_ifbz: {first_column}')

second_column = {}
for class_name in classesList:
    second_column[class_name] = {}
    second_column[class_name] = copy.deepcopy(first_column[class_name])
    for feature in featuresList:
        second_column[class_name][feature] = {}
        second_column[class_name][feature] = copy.deepcopy(first_column[class_name][feature])
        for medical_history in medicalHistoryList:
            second_column[class_name][feature][medical_history] = {}
            fc = copy.deepcopy(list(first_column[class_name][feature][medical_history].items()))
            pair = []
            for i in range(len(fc) - 1):
                pair.append((fc[i][0], fc[i + 1][0]))
            alternatives = {}
            a_i = 0
            while a_i < len(fc):
                a_i += 1
                if a_i == 1:
                    alternatives[f'Альтернатива {1}.{1}'] = fc[len(fc) - 1][0]
                    continue
                else:
                    combinations = list(itertools.combinations(pair, a_i - 1))
                    for a_j in range(len(combinations)):
                        # todo: исправить последнюю единицу на 0, после дебага
                        if a_i == 2:
                            alternatives[f'Альтернатива {a_i}.{a_j + 1}'] = [pair[a_j], fc[len(fc) - 1][0]]
                        else:
                            alternatives[f'Альтернатива {a_i}.{a_j + 1}'] = [combinations[a_j], fc[len(fc) - 1][0]]
            second_column[class_name][feature][medical_history] = alternatives

third_column = {}
for class_name in classesList:
    third_column[class_name] = {}
    third_column[class_name] = copy.deepcopy(second_column[class_name])
    for feature in featuresList:
        third_column[class_name][feature] = {}
        third_column[class_name][feature] = copy.deepcopy(second_column[class_name][feature])
        for medical_history in medicalHistoryList:
            third_column[class_name][feature][medical_history] = {}
            alternatives = {}
            list_fc_keys = list(copy.deepcopy(first_column[class_name][feature][medical_history]).keys())
            list_fc_value = list(copy.deepcopy(first_column[class_name][feature][medical_history]).values())
            list_fc_items = list(copy.deepcopy(first_column[class_name][feature][medical_history]).items())
            this_PD = 1
            counter_true_alternative = 1
            for key, value in copy.deepcopy(second_column[class_name][feature][medical_history]).items():
                map_for_alternatives = {}
                num_PD = int(key[key.index(' ') + 1:key.index(
                    '.')])  # Приходится искать точку и пробел для того чтобы получить правильный номер ПД
                if num_PD != this_PD:
                    this_PD = num_PD
                    counter_true_alternative = 1
                if key == 'Альтернатива 1.1':
                    # убирет повторы для каждого здп
                    map_for_alternatives['ЗДП'] = [set(list_fc_value)]
                    map_for_alternatives['ВГ'] = [value]
                    map_for_alternatives['НГ'] = [value]
                    alternatives[key] = map_for_alternatives
                    continue
                else:
                    flag_for_true_alternative = True
                    val = value[:-1]
                    list_index_val, list_set_val, list_border = [], [], []
                    buf = 0
                    border = 0
                    if type(val[0][0]) is tuple:
                        val = val[0]
                    for i in range(len(val)):
                        key_1, key_2 = val[i]
                        list_border.append(key_1 - border)
                        border = (key_1 + key_2) // 2  # Граница между периодами
                        list_index_val.append([buf, list_fc_keys.index(key_1) + 1])  # Точно + 1?
                        buf = list_fc_keys.index(key_2)
                        if i == len(val) - 1:
                            list_index_val.append([buf, len(list_fc_value)])
                            list_border.append(list_fc_keys[-1] - border)
                    for i in range(len(list_index_val)):
                        ind_1, ind_2 = list_index_val[i]
                        list_set_val.append(set(list_fc_value[ind_1: ind_2]))
                    for i in range(1, len(list_set_val)):
                        mas_1 = list_set_val[i - 1]
                        mas_2 = list_set_val[i]
                        if mas_1.intersection(mas_2):
                            flag_for_true_alternative = False
                            break
                    if flag_for_true_alternative:
                        map_for_alternatives['ЗДП'] = copy.deepcopy(list_set_val)
                        map_for_alternatives['ВГ'] = copy.deepcopy(list_border)
                        map_for_alternatives['НГ'] = copy.deepcopy(list_border)
                        key_for_true_alternative = f'Альтернатива {this_PD}.{counter_true_alternative}'
                        alternatives[
                            key_for_true_alternative] = map_for_alternatives  # можно подставить key, если хочется видеть пробелы в нейминге альтернатив
                        counter_true_alternative += 1
            third_column[class_name][feature][medical_history] = alternatives

fifth_column = {}
variant_num = 1
variant_num_2 = 1
for class_name in classesList:
    fifth_column[class_name] = {}
    fifth_column[class_name] = copy.deepcopy(second_column[class_name])
    for feature in featuresList:
        fifth_column[class_name][feature] = {}
        alternatives = {}
        IB_alter = {}
        for pd_generation in range(1, 10):
            alternatives[pd_generation] = {}
            alternatives[pd_generation]["ЗДП"] = []
            alternatives[pd_generation]["ВГ"] = []
            alternatives[pd_generation]["НГ"] = []
        for key, value in copy.deepcopy(third_column[class_name][feature][medicalHistoryList[0]]).items():
            num_PD = int(key[key.index(' ') + 1:key.index('.')])
            alternatives[num_PD]["ЗДП"].append(value["ЗДП"])
            alternatives[num_PD]["ВГ"].append(value["ВГ"])
            alternatives[num_PD]["НГ"].append(value["НГ"])
        for medical_history in medicalHistoryList[1:]:  # Первая история болезни записывается полностью
            IB_alter = {}
            for pd_generation in range(1, 10):
                IB_alter[pd_generation] = {}
                IB_alter[pd_generation]["ЗДП"] = []
                IB_alter[pd_generation]["ВГ"] = []
                IB_alter[pd_generation]["НГ"] = []
            for key, value in copy.deepcopy(third_column[class_name][feature][medical_history]).items():
                num_PD = int(key[key.index(' ') + 1:key.index('.')])
                IB_alter[num_PD]["ЗДП"].append(value["ЗДП"])
                IB_alter[num_PD]["ВГ"].append(value["ВГ"])
                IB_alter[num_PD]["НГ"].append(value["НГ"])
            alternatives[1]["ЗДП"][0][0].update(IB_alter[1]["ЗДП"][0][0])
            alternatives[1]["ВГ"][0][0] = max(IB_alter[1]["ВГ"][0][0], alternatives[1]["ВГ"][0][0])
            alternatives[1]["НГ"][0][0] = min(IB_alter[1]["НГ"][0][0], alternatives[1]["НГ"][0][0])
            for pd_iter in range(2, 10):
                ZDP_List = []
                VG_List = []
                NG_List = []
                for i in range(len(alternatives[pd_iter]["ЗДП"])):
                    for j in range(len(IB_alter[pd_iter]["ЗДП"])):
                        flag = True
                        list_iter_alternatives, vg_alternatives, ng_alternatives = [], [], []
                        for iter_PD_from_alternatives in range(len(alternatives[pd_iter]["ЗДП"][i])):
                            list_iter_alternatives.append(
                                alternatives[pd_iter]["ЗДП"][i][iter_PD_from_alternatives].union(
                                    IB_alter[pd_iter]["ЗДП"][j][iter_PD_from_alternatives]))
                            vg_alternatives.append(max(IB_alter[pd_iter]["ВГ"][j][iter_PD_from_alternatives],
                                                       alternatives[pd_iter]["ВГ"][i][iter_PD_from_alternatives]))
                            ng_alternatives.append(min(IB_alter[pd_iter]["НГ"][j][iter_PD_from_alternatives],
                                                       alternatives[pd_iter]["НГ"][i][iter_PD_from_alternatives]))
                        for iter_PD_from_alternatives in range(1, len(list_iter_alternatives)):
                            mas_1 = list_iter_alternatives[iter_PD_from_alternatives - 1]
                            mas_2 = list_iter_alternatives[iter_PD_from_alternatives]
                            if mas_1.intersection(mas_2):
                                flag = False
                                break
                        if flag:
                            ZDP_List.append(list_iter_alternatives)
                            VG_List.append(vg_alternatives)
                            NG_List.append(ng_alternatives)
                alternatives[pd_iter]["ЗДП"] = copy.deepcopy(ZDP_List)
                alternatives[pd_iter]["ВГ"] = copy.deepcopy(VG_List)
                alternatives[pd_iter]["НГ"] = copy.deepcopy(NG_List)
        fifth_column[class_name][feature] = copy.deepcopy(alternatives)
# print(fifth_column)
# print(improvedAmountPeriodDynamicsForFeaturesInClassMap)

last_column = {}
for class_name in classesList:
    last_column[class_name] = {}
    for feature in featuresList:
        counter_all_alternative = 0
        counter_true_alternative = 0
        row_ter = 1
        last_column[class_name][feature] = {}
        for pd_iter in range(1, 10):
            size_pd = len(fifth_column[class_name][feature][pd_iter]['ЗДП'])
            if size_pd == 0:
                continue
            counter_all_alternative += size_pd
            if len(improvedAmountPeriodDynamicsForFeaturesInClassMap[class_name][feature].keys()) == pd_iter:
                counter_true_alternative += size_pd
                for iter_alternative in range(len(fifth_column[class_name][feature][pd_iter]['ЗДП'])):
                    for iter_from_pd in range(pd_iter):
                        last_column[class_name][feature][row_ter] = {}
                        last_column[class_name][feature][row_ter]["Класс"] = class_name.split()[1]
                        last_column[class_name][feature][row_ter]["Признак"] = feature.split()[1]
                        last_column[class_name][feature][row_ter]["ЧПД"] = pd_iter
                        last_column[class_name][feature][row_ter]["ПД"] = iter_from_pd + 1
                        if type(improvedAmountPeriodDynamicsForFeaturesInClassMap[class_name][feature][iter_from_pd + 1]) is tuple:
                            left, right = improvedAmountPeriodDynamicsForFeaturesInClassMap[class_name][feature][iter_from_pd + 1]
                            last_column[class_name][feature][row_ter]["ЗДП МБЗ"] = "[" + str(left) + ", " + str(right) + "]"
                        elif type(improvedAmountPeriodDynamicsForFeaturesInClassMap[class_name][feature][iter_from_pd + 1]) is int:
                            last_column[class_name][feature][row_ter]["ЗДП МБЗ"] = improvedAmountPeriodDynamicsForFeaturesInClassMap[class_name][feature][iter_from_pd + 1]
                        else:
                            last_column[class_name][feature][row_ter]["ЗДП МБЗ"] = ', '.join(improvedAmountPeriodDynamicsForFeaturesInClassMap[class_name][feature][iter_from_pd + 1])
                        last_column[class_name][feature][row_ter]["ЗДП ИФБЗ"] = fifth_column[class_name][feature][pd_iter]['ЗДП'][iter_alternative][iter_from_pd]
                        ngmbz, vgmbz = improvedAmountPeriodDynamicsForFeaturesInClassVGNGMap[class_name][feature][iter_from_pd + 1]
                        ngIfbz = fifth_column[class_name][feature][pd_iter]['НГ'][iter_alternative][iter_from_pd]
                        vgIfbz = fifth_column[class_name][feature][pd_iter]['ВГ'][iter_alternative][iter_from_pd]
                        last_column[class_name][feature][row_ter]["ВГ"] = vgmbz - vgIfbz
                        last_column[class_name][feature][row_ter]["НГ"] = ngmbz - ngIfbz
                        row_ter += 1

# ----------------------------------------------- Вывод -----------------------------------------------
# Создаем новый файл
workbook = Workbook()

# Создаем 4 страницы
workbook.create_sheet("МБЗ")
workbook.create_sheet("МВД")
workbook.create_sheet("ИФБЗ")
workbook.create_sheet("МБЗ vs. ИФБЗ")

# Получаем страницы
mbz = workbook["МБЗ"]
mvd = workbook["МВД"]
ifbz = workbook["ИФБЗ"]
vs = workbook["МБЗ vs. ИФБЗ"]
# workbook.remove_sheet(workbook['Sheet'])

# Задаем стиль ячейки для жирного шрифта
boldFont = Font(bold=True)

# Задаем стиль ячейки для заливки цветом
fillGreen = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

# Задаем стиль ячейки для обводки
thinBorder = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
# ----------------------------------------------- МБЗ -----------------------------------------------
# Классы
mbz.cell(row=1, column=1).value = "Классы"
mbz.cell(row=1, column=1).font = boldFont
for i in range(len(classesList)):
    cell = mbz.cell(row=i + 2, column=1, value=classesList[i])

# Признаки
mbz.cell(row=1, column=3).value = "Признаки"
mbz.cell(row=1, column=3).font = boldFont
for i in range(len(featuresList)):
    cell = mbz.cell(row=i + 2, column=3, value=featuresList[i])

# Возможные значения (ВЗ)
mbz.cell(row=1, column=5).value = "Возможные значения (ВЗ)"
mbz.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)
mbz.cell(row=1, column=5).font = boldFont
for i in range(len(featuresList)):
    cellLeft = mbz.cell(row=i + 2, column=5, value=featuresList[i])
    cellRight = mbz.cell(row=i + 2, column=6, value=str(possibleValuesForFeaturesMap[featuresList[i]]))

# Нормальные значения (НЗ)
mbz.cell(row=1, column=8).value = "Нормальные значения (НЗ)"
mbz.merge_cells(start_row=1, start_column=8, end_row=1, end_column=9)
mbz.cell(row=1, column=8).font = boldFont
for i in range(len(featuresList)):
    cellLeft = mbz.cell(row=i + 2, column=8, value=featuresList[i])
    cellRight = mbz.cell(row=i + 2, column=9, value=str(improvedMap[featuresList[i]][1]))

# Клиническая картина (КК)
mbz.cell(row=1, column=11).value = "Клиническая картина (КК)"
mbz.merge_cells(start_row=1, start_column=11, end_row=1, end_column=12)
mbz.cell(row=1, column=11).font = boldFont
for j in range(len(classesList)):
    for i in range(len(featuresList)):
        cellLeft = mbz.cell(row=i + 2 + (COUNT_FEATURE_INTO_CLASS * j), column=11, value=classesList[j])
        cellRight = mbz.cell(row=i + 2 + (COUNT_FEATURE_INTO_CLASS * j), column=12, value=featuresList[i])

# Число периодов динамики (ЧПД)
mbz.cell(row=1, column=14).value = "Число периодов динамики (ЧПД)"
mbz.merge_cells(start_row=1, start_column=14, end_row=1, end_column=16)
mbz.cell(row=1, column=14).font = boldFont
for j in range(len(classesList)):
    for i in range(len(featuresList)):
        CHPD = amountPeriodDynamicsForFeaturesInClassMap[classesList[j]][featuresList[i]]
        cellLeft = mbz.cell(row=i + 2 + (COUNT_FEATURE_INTO_CLASS * j), column=14, value=classesList[j])
        cellMiddle = mbz.cell(row=i + 2 + (COUNT_FEATURE_INTO_CLASS * j), column=15, value=featuresList[i])
        cellRight = mbz.cell(row=i + 2 + (COUNT_FEATURE_INTO_CLASS * j), column=16, value=CHPD)

# Значения для периода (ЗДП)
mbz.cell(row=1, column=18).value = "Значения для периода (ЗДП)"
mbz.merge_cells(start_row=1, start_column=18, end_row=1, end_column=21)
mbz.cell(row=1, column=18).font = boldFont
count: int = 0
for class_name in classesList:
    for i, feature in enumerate(featuresList):
        CHPD = amountPeriodDynamicsForFeaturesInClassMap[class_name][feature]
        for k in range(1, CHPD + 1):
            ZDP = str(improvedAmountPeriodDynamicsForFeaturesInClassMap[class_name][feature][k])
            row = i + 2 + count
            cellLeft = mbz.cell(row=row, column=18, value=class_name)
            cellMiddleLeft = mbz.cell(row=row, column=19, value=feature)
            cellMiddleRight = mbz.cell(row=row, column=20, value=k)
            cellRight = mbz.cell(row=row, column=21, value=ZDP)
            # print(f'cellLeft: {cellLeft.value}, cellMiddleLeft: {cellMiddleLeft.value}, cellRight: {cellMiddleRight.value}, cellMiddleRight: {cellRight.value}')
            if k != CHPD:
                count += 1
    count += COUNT_CHPD_END + 1

# Верхние и нижние границы (ВГ и НГ)
mbz.cell(row=1, column=23).value = "Верхние и нижние границы (ВГ и НГ)"
mbz.merge_cells(start_row=1, start_column=23, end_row=1, end_column=27)
mbz.cell(row=1, column=23).font = boldFont
count: int = 0
for class_name in classesList:
    for i, feature in enumerate(featuresList):
        CHPD = amountPeriodDynamicsForFeaturesInClassMap[class_name][feature]
        for k in range(1, CHPD + 1):
            VGNG = improvedAmountPeriodDynamicsForFeaturesInClassVGNGMap[class_name][feature][k]
            row = i + 2 + count
            cellLeft = mbz.cell(row=row, column=23, value=class_name)
            cellMiddleLeft = mbz.cell(row=row, column=24, value=feature)
            cellMiddleMiddle = mbz.cell(row=row, column=25, value=k)
            cellMiddleRight = mbz.cell(row=row, column=26, value=VGNG[1])
            cellRight = mbz.cell(row=row, column=27, value=VGNG[0])
            if k != CHPD:
                count += 1
    count += COUNT_CHPD_END + 1
# ----------------------------------------------- МБЗ -----------------------------------------------
# (ИБ, заболевание, признак, номер ПД, длительность ПД, число МН в ПД)
mvd.cell(row=1, column=1).value = "(ИБ, заболевание, признак, номер ПД, длительность ПД, число МН в ПД)"
mvd.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
mvd.cell(row=1, column=1).font = boldFont

count: int = 0
for medical_history in medicalHistoryList:
    for class_name in classesList:
        for i, feature in enumerate(featuresList):
            CHPD = amountPeriodDynamicsForFeaturesInClassMap[class_name][feature]
            for k in range(1, CHPD + 1):
                MVD = medicalHistoryMap[medical_history][class_name][feature][k]
                row = i + 2 + count
                cell_1 = mvd.cell(row=row, column=1, value=medical_history)
                cell_2 = mvd.cell(row=row, column=2, value=class_name)
                cell_3 = mvd.cell(row=row, column=3, value=feature)
                cell_4 = mvd.cell(row=row, column=4, value=k)
                cell_5 = mvd.cell(row=row, column=5, value=MVD['Длительность ПД'])
                cell_6 = mvd.cell(row=row, column=6, value=MVD['Число МН в ПД'])
                if k != CHPD:
                    count += 1
        count += COUNT_CHPD_END + 1

# (ИБ, заболевание, признак, номер ПД, длительность ПД, число МН в ПД)
mvd.cell(row=1, column=8).value = "Выборка данных (ИБ, заболевание, признак, МН, значение в МН)"
mvd.merge_cells(start_row=1, start_column=8, end_row=1, end_column=12)
mvd.cell(row=1, column=8).font = boldFont

count: int = 0
for medical_history in medicalHistoryList:
    for class_name in classesList:
        for i, feature in enumerate(featuresList):
            for k, data in enumerate(data_sampling[medical_history][class_name][feature].items()):
                row = i + 2 + count
                cell_1 = mvd.cell(row=row, column=8, value=medical_history)
                cell_2 = mvd.cell(row=row, column=9, value=class_name)
                cell_3 = mvd.cell(row=row, column=10, value=feature)
                cell_4 = mvd.cell(row=row, column=11, value=str(f'МН({data[0]})'))
                cell_5 = mvd.cell(row=row, column=12, value=data[1])
                if k != len(data_sampling[medical_history][class_name][feature].items()) - 1:
                    count += 1
        count += COUNT_CHPD_END + 1
# ----------------------------------------------- ИФБЗ -----------------------------------------------
import json

# Запись переменной first_column в файл
with open('ifbz_first_column.txt', 'w') as file:
    for class_name in classesList:
        print(f'{class_name}', file=file)
        for feature in featuresList:
            print(f'  {feature}', file=file)
            for medical_history in medicalHistoryList:
                print(f'    {medical_history}', file=file)
                for key, value in first_column[class_name][feature][medical_history].items():
                    print(f'      {key} - {value}', file=file)

# Запись переменной second_column в файл
with open('ifbz_second_column.txt', 'w') as file:
    for class_name in classesList:
        print(f'{class_name}', file=file)
        for feature in featuresList:
            print(f'  {feature}', file=file)
            for medical_history in medicalHistoryList:
                print(f'    {medical_history}', file=file)
                for key, value in second_column[class_name][feature][medical_history].items():
                    print(f'      {key} - {value}', file=file)

# Запись переменной third_column в файл
with open('ifbz_third_column.txt', 'w') as file:
    for class_name in classesList:
        print(f'{class_name}', file=file)
        for feature in featuresList:
            print(f'  {feature}', file=file)
            for medical_history in medicalHistoryList:
                print(f'    {medical_history}', file=file)
                for name_alternative, value_alternative in third_column[class_name][feature][medical_history].items():
                    print(f'      {name_alternative}', file=file)
                    for zdp, vg, ng in zip(third_column[class_name][feature][medical_history][name_alternative]['ЗДП'], third_column[class_name][feature][medical_history][name_alternative]['ВГ'], third_column[class_name][feature][medical_history][name_alternative]['НГ']):
                        print(f'        ЗДП: {zdp} - ВГ: {vg} - НГ: {ng}', file=file)
                    

# Запись переменной fifth_column в файл
with open('ifbz_fifth_column.txt', 'w') as file:
    for class_name in classesList:
        print(f'{class_name}', file=file)
        for feature in featuresList:
            print(f'  {feature}', file=file)
            for key, value in fifth_column[class_name][feature].items():
                if value['ЗДП'] == []:
                    continue
                else:
                    print(f'    ЗДП с {key} периодом(-ов/-ами) динамики', file=file)
                    for zdp, vg, ng in zip(value['ЗДП'], value['ВГ'], value['НГ']):
                        print(f'      ЗДП: {zdp} - ВГ: {vg} - НГ: {ng}', file=file)
# ----------------------------------------------- VS -----------------------------------------------

# ----------------------------------------------- OVER -----------------------------------------------
# проходим по всем листам книги
for worksheet in workbook.worksheets:
    # проходим по всем строкам и столбцам листа
    for row in worksheet.iter_rows():
        for cell in row:
            # Проверяем, есть ли значение в ячейке
            if cell.value:
                # Применяем стили
                cell.border = thinBorder
                cell.fill = fillGreen

for worksheet in workbook.worksheets:
    # Проходим по всем столбцам
    for column in worksheet.columns:
        # Инициализируем переменные для объединения
        start_cell = column[0]
        end_cell = column[0]
        previous_value = column[0].value
        # Проходим по всем ячейкам в столбце, начиная со второй
        for cell in column[1:]:
            # Если значение ячейки совпадает с предыдущей ячейкой
            if cell.value == previous_value:
                # Обновляем переменную для конечной ячейки
                end_cell = cell
            else:
                # Если значение ячейки не совпадает с предыдущей ячейкой,
                # то объединяем ячейки от start_cell до end_cell
                if start_cell != end_cell:
                    worksheet.merge_cells(start_row=start_cell.row, start_column=start_cell.column,
                                          end_row=end_cell.row, end_column=end_cell.column)
                # Обновляем переменные для начальной и конечной ячеек
                start_cell = cell
                end_cell = cell
                previous_value = cell.value

        # Объединяем последнюю группу ячеек в столбце
        if start_cell != end_cell:
            worksheet.merge_cells(start_row=start_cell.row, start_column=start_cell.column,
                                  end_row=end_cell.row, end_column=end_cell.column)

# Сохраняем файл
workbook.save(filename="IAD.xlsx")

print(last_column)